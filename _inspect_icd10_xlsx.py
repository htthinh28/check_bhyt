# -*- coding: utf-8 -*-
import json
import sys
sys.stdout.reconfigure(encoding='utf-8')
import openpyxl

p = r"c:\Users\admin\Downloads\Phu luc Bang danh muc ICD10_FINAL  (1).xlsx"
wb = openpyxl.load_workbook(p, read_only=True, data_only=True)
print("sheets:", wb.sheetnames)
for sn in wb.sheetnames:
    ws = wb[sn]
    rows = list(ws.iter_rows(max_row=3, values_only=True))
    print("\n===", sn, "cols:", len(rows[0]) if rows else 0)
    for i, r in enumerate(rows):
        print(i, list(r)[:25])
