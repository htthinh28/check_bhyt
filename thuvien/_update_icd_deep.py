# -*- coding: utf-8 -*-
"""
Cập nhật gợi ý ICD-10: mã chi tiết, tách 4 cột mã/tên (chỉ định & chống chỉ định).
Chạy: python _update_icd_deep.py [--dry-run]
"""
from __future__ import annotations

import argparse
import json
import re
import sys
import unicodedata
from collections import defaultdict
from pathlib import Path

sys.stdout.reconfigure(encoding="utf-8")

WORKSPACE = Path(__file__).resolve().parent
HTML_GLOB = "*(1)*.html"
XLSX_DEFAULT = r"c:\Users\admin\Downloads\Phu luc Bang danh muc ICD10_FINAL  (1).xlsx"

ICD_RE = re.compile(r"^([A-TV-ZU]\d{2}(?:\.\d{1,2})?\*?)")

PHRASE_HINTS: list[tuple[list[str], str]] = [
    (["nhiem hiv", " hiv", "aids", "suy giam mien dich"], "B24"),
    (["viem tai giua", "otitis media"], "H66"),
    (["viem xoang", "sinusitis"], "J01"),
    (["viem amidan", "viem amydan", "viem hong", "viem amiđan"], "J03"),
    (["viem phoi mac phai", "viem phoi cap tai cong dong"], "J18"),
    (["viem phoi benh vien"], "J15"),
    (["viem phoi", "pneumonia"], "J18"),
    (["viem phe quan man", "dot bung phat", "copd"], "J44"),
    (["nhiem trung da", "nhiem khuan da", "viem da va mo", "ap xe rang", "viem mo te bao lan toa"], "L03"),
    (["viem bang quang cap"], "N30"),
    (["viem bang quang"], "N30"),
    (["viem than", "be than", "viem than-be than"], "N10"),
    (["loet da day", "loet da day-ta trang", "helicobacter", "hpylori"], "K25"),
    (["benh lyme", "lyme"], "A69.2"),
    (["anthrax", "bacillus anthracis", "benh than do vi khuan", "phoi nhiem than", "benh than do"], "A22"),
    (["benh lau", "gonorrh", "gonococc"], "A54"),
    (["viem mang nao", "meningitis"], "G03"),
    (["tang huyet ap", "tăng huyết áp", "tăng huyết áp nguyên phát"], "I10"),
    (["suy tim mat bu", "suy tim sung huyet"], "I50"),
    (["suy tim", "suy thất trái", "suy that trai"], "I50"),
    (["nhoi mau co tim", "nhiem mau co tim", "st chênh lên", "st chech len"], "I21"),
    (["dot quy", "thieu mau cuc bo", "cơn thiếu máu cục bộ"], "I63"),
    (["dai thao duong typ 2", "dtd typ 2", "dtd type 2", "diabetes typ 2", "đái tháo đường typ 2", "dai thao duong type 2"], "E11"),
    (["dai thao duong", "đái tháo đường", "duong huyet tang", "duong huyet cao", "tieu duong"], "E11"),
    (["ung thu vu", "ung thư vú", "carcinoma vu"], "C50"),
    (["ung thu phoi", "ung thư phổi", "carcinoma phoi"], "C34"),
    (["ung thu tuyen tien liet", "ung thư tuyến tiền liệt", "prostate"], "C61"),
    (["ung thu da day", "ung thư dạ dày", "adenocarcinoma da day"], "C16"),
    (["ung thu dau", "ung thư đầu", "ung thu co", "đầu - cổ"], "C00"),
    (["chung dau", "chứng đau", "giam dau", "giảm đau", " dau nhe", " đau vua"], "R52"),
    ([" ha sot", "hạ sốt", " sot "], "R50"),
    (["suy gan nang", "suy gan vua"], "K72"),
    (["suy gan"], "K72"),
    (["suy than", "suy thận", "gfr", "loc cau than"], "N18"),
    ([" lao", "lao:", "tuberculosis", "mycobacterium tuberculosis"], "A15"),
    (["dich hai", "peste", "yersinia pestis"], "A20"),
    (["brucella"], "A23"),
    (["sot vang", "yellow fever"], "A95"),
    (["viem khop", "nhiem khuan xuong", "nhiễm khuẩn xương"], "M00"),
    (["viem mang bung", "viêm màng bụng"], "K65"),
    (["nhiem khuan huyet", "sepsis", "sot giam bach cau", "nhiễm khuẩn huyết"], "A41"),
    (["hen ", " hen", "asthma", "hen phe quan"], "J45"),
    (["viem tai ngoai", "viêm tai ngoài"], "H60"),
    (["viem phe quan cap", "viêm phế quản cấp"], "J20"),
    (["benh than o", "benh than dai thao duong", "bệnh thận đái tháo đường", "protein nieu"], "N08"),
    (["co giat", "co giật", "epilep"], "G40"),
    (["dau than kinh", "đau thần kinh", "neuropathic"], "G62"),
    (["benh parkinson", "parkinson"], "G20"),
    (["rung nhi", "rung nhĩ", "fibrillation"], "I48"),
    (["viem loet dai trang", "colitis", "viêm loét đại tràng"], "K51"),
    (["viem da day", "viêm dạ dày"], "K29"),
    (["tang lipid", "tăng lipid", "cholesterol", "rối loạn lipid"], "E78"),
    (["basedow", "cường giáp", "cuong giap"], "E05"),
    (["suy giap", "suy giáp"], "E03"),
    (["viem gan b", "viêm gan b", "hepatitis b"], "B18"),
    (["viem gan c", "viêm gan c", "hepatitis c"], "B18.2"),
    (["benh giang mai", "giang mai", "syphilis"], "A53"),
    (["nong do benzodiazepin", "ngộ độc benzodiazepin", "ngu doc benzodiazepin"], "T42"),
    (["ngu doc", "ngộ độc", "qua lieu"], "T50"),
    (["teo day than kinh thi giac leber"], "H47"),
    (["giam thi luc do thuoc la"], "H53"),
    (["suy tim cung luong cao", "soc nhiem khuan"], "R57"),
    (["hep eo dong mach chu", "hẹp eo động mạch chủ"], "Q25"),
    (["benh da hong cau", "đa hồng cầu"], "D45"),
    (["bach cau tuy bao man", "bạch cầu tủy bào mạn"], "C92"),
    (["xo hoa tuy", "xơ hóa tủy"], "D75"),
    (["tang tieu cau", "tăng tiểu cầu"], "D47"),
    (["benh te bao goc", "tế bào gốc"], "Z51"),
    (["viem khop dang thap", "viêm khớp dạng thấp"], "M05"),
    (["thap khop", "thấp khớp"], "M06"),
    (["gout", "gút"], "M10"),
    (["loang xuong", "loãng xương"], "M81"),
    (["benh nhiem trung duong sinh duc", "herpes simplex"], "A60"),
    (["viem gan man", "viêm gan mạn"], "K73"),
]

MULTI_DEFAULT: dict[str, list[str]] = {
    "E11": ["E11.8", "E11.9"],
    "J18": ["J18.8", "J18.9"],
    "J01": ["J01.9"],
    "J03": ["J03.9"],
    "H66": ["H66.9"],
    "G03": ["G03.9"],
    "I63": ["I63.9"],
    "K72": ["K72.9"],
    "N18": ["N18.9"],
    "T78": ["T78.9"],
    "A15": ["A15.9"],
    "N30": ["N30.0", "N30.9"],
    "J45": ["J45.9"],
    "I50": ["I50.9"],
    "I48": ["I48.9"],
    "R52": ["R52.9"],
    "R50": ["R50.9"],
    "C50": ["C50.9"],
    "I21": ["I21.9"],
}

# Chỉ dùng cho chongChiDinh — không map bệnh điều trị (E11, I10, J18…)
CHONG_PHRASE_HINTS: list[tuple[list[str], str]] = [
    (["qua man", "mẫn cảm với", "mẫn cảm", "di ung", "dị ứng", "di ung voi", "dị ứng với", "tien su qua man", "tiền sử quá mẫn", "phan ung di ung"], "T78"),
    (["mang thai", "phu nu mang thai", "nguoi mang thai", "thai ky", "thai kỳ", "trong thai ky", "3 thang giua", "3 thang cuoi"], "Z33"),
    (["suy gan nang", "suy gan vua", "suy gan"], "K72"),
    (["suy than", "suy thận", "suy than man", "gfr", "loc cau than", "muc loc cau than"], "N18"),
    (["suy tim cap", "suy tim chua kiem soat", "suy tim mat bu cap", "soc tim", "suy tim do iii", "suy tim do iv"], "I50"),
    (["bloc nhi that", "block nhi that", "block av"], "I44"),
    (["chay mau dang hoat dong", "benh chay mau", "nguy co chay mau co y nghia"], "D68"),
    (["nhiem trung toan than dang hoat dong", "nhiem trung toan than"], "A49"),
    (["hen cap", "con hen cap", "cat con hen cap"], "J45"),
    (["viem gan cap", "suy chuc nang gan nang"], "K72"),
    (["mang thai 3 thang"], "Z33"),
]

# Mệnh đề mô tả điều kiện BN / phối hợp thuốc — không phải mã ICD chống chỉ định
RELATIVE_CONTRA_CLAUSE = re.compile(
    r"phoi hop .{0,120} chong chi dinh tren|"
    r"chong chi dinh tren benh nhan|"
    r"la chong chi dinh tren|"
    r"tren benh nhan (bi |co )?dai thao duong|"
    r"benh nhan bi dai thao duong|"
    r"benh nhan (bi )?dtd|"
    r"hoac suy than\s*\(|"
    r"khi phoi hop",
    re.I,
)

CHONG_CLAUSE_SIGNAL = re.compile(
    r"chong chi dinh|khong (duoc |nen )?dung|cam dung|tranh dung|ngung dung|"
    r"qua man|mau cam|di ung|mang thai|suy gan|suy than|suy tim|chay mau|"
    r"bloc nhi|hen cap|nhiem trung toan than",
    re.I,
)

# Không map vào chỉ định (chỉ dùng cho chống chỉ định)
CHI_EXCLUDED_TARGETS = frozenset({"T78", "Z33"})


def norm(s: str) -> str:
    s = unicodedata.normalize("NFD", str(s or ""))
    s = "".join(c for c in s if unicodedata.category(c) != "Mn")
    s = s.replace("đ", "d").replace("Đ", "D")
    s = re.sub(r"[^\w\s.*]", " ", s.lower())
    s = re.sub(r"\s+", " ", s).strip()
    return s


def split_clauses(text: str) -> list[str]:
    if not text:
        return []
    t = str(text).replace("\r\n", "\n")
    parts: list[str] = []
    for line in re.split(r"[\n;]+", t):
        line = re.sub(r"^[•\-\*·▪–—]\s*", "", line.strip())
        line = re.sub(r"^\d+[.)]\s*", "", line)
        if line:
            parts.append(line)
    return parts


def score_match(clause_norm: str, entry: dict) -> float:
    ten = entry["norm_ten"]
    if not ten or len(ten) < 6:
        return 0.0
    if ten in clause_norm:
        return 200.0 + len(ten)
    words = [w for w in ten.split() if len(w) >= 5]
    hits = sum(1 for w in words if w in clause_norm)
    if hits >= 2:
        return 90.0 + hits * 8
    return 0.0


def pick_best_under_prefix(prefix: str, clause_norm: str, catalog: dict, children: dict) -> dict | None:
    prefix = prefix.upper().rstrip("*")
    if prefix in catalog and "." in prefix:
        return catalog[prefix]
    base = prefix.split(".")[0]
    cands = children.get(base, [])
    dotted = [c for c in cands if "." in c["ma"]]
    if not dotted:
        return catalog.get(prefix) or catalog.get(base)
    best = None
    best_sc = 0.0
    for c in dotted:
        sc = score_match(clause_norm, c)
        if sc > best_sc:
            best_sc = sc
            best = c
    if best and best_sc >= 35:
        return best
    for c in dotted:
        if c["ma"].endswith(".9"):
            return c
    return dotted[0] if dotted else catalog.get(base)


def resolve_targets(target: str, clause_norm: str, catalog: dict, children: dict) -> list[dict]:
    target = target.upper().rstrip("*")
    base = target.split(".")[0]

    if base in MULTI_DEFAULT and target in (base, f"{base}."):
        out = [catalog[c] for c in MULTI_DEFAULT[base] if c in catalog]
        if out:
            return out

    if target in catalog and "." in target:
        return [catalog[target]]

    picked = pick_best_under_prefix(target, clause_norm, catalog, children)
    if not picked:
        return []

    ma = picked["ma"]
    b = ma.split(".")[0]
    dotted = [c for c in children.get(b, []) if "." in c["ma"]]

    if "." not in ma and (picked.get("flag4") or len(dotted) > 1):
        if b in MULTI_DEFAULT:
            return [catalog[c] for c in MULTI_DEFAULT[b] if c in catalog]
        refined = pick_best_under_prefix(b, clause_norm, catalog, children)
        return [refined] if refined else [picked]

    if ma == base and b in MULTI_DEFAULT:
        return [catalog[c] for c in MULTI_DEFAULT[b] if c in catalog]

    return [picked]


def load_icd_from_html(html_text: str):
    catalog: dict[str, dict] = {}
    children: dict[str, list[dict]] = defaultdict(list)
    for m in re.finditer(r'class="icd-data-chunk"[^>]*>(\[.*?\])</script>', html_text, re.S):
        for x in json.loads(m.group(1)):
            ma = str(x.get("m") or "").strip().upper()
            if not ma or not ICD_RE.match(ma):
                continue
            ten = str(x.get("t") or "").strip()
            flag4 = "4" in (x.get("f") or [])
            e = {"ma": ma, "ten": ten, "norm_ten": norm(ten), "flag4": flag4}
            catalog[ma] = e
            children[ma.split(".")[0]].append(e)
    return catalog, list(catalog.values()), children


def try_load_xlsx(xlsx_path: str):
    try:
        import openpyxl
    except ImportError:
        return None
    p = Path(xlsx_path)
    if not p.is_file():
        return None
    wb = openpyxl.load_workbook(p, read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]
    catalog: dict[str, dict] = {}
    children: dict[str, list[dict]] = defaultdict(list)
    for row in ws.iter_rows(min_row=5, values_only=True):
        ma = str(row[17] or "").strip().upper()
        if not ma or not ICD_RE.match(ma):
            continue
        ten = str(row[21] or "").strip()
        flag4 = bool(str(row[25] or "").strip())
        e = {"ma": ma, "ten": ten, "norm_ten": norm(ten), "flag4": flag4}
        catalog[ma] = e
        children[ma.split(".")[0]].append(e)
    return catalog, list(catalog.values()), children


def any_hint_match(cn: str, patterns: list[str]) -> bool:
    return any(norm(p) in cn for p in patterns)


def map_text_to_icds(text: str, catalog: dict, children: dict, max_codes: int = 25) -> list[dict]:
    results: dict[str, dict] = {}
    clauses = split_clauses(text)
    norms = [norm(c) for c in clauses if c.strip()]
    if not norms:
        norms = [norm(text)]

    work_norm = " ".join(norms)
    if re.search(r"khong (co|duoc) (tac dung|dung|chi dinh)", work_norm):
        for seg in re.split(r"[.;:\n]", work_norm):
            if re.search(r"khong (co|duoc) (tac dung|dung|chi dinh)", seg):
                work_norm = work_norm.replace(seg, " ")

    search_norms = norms + [work_norm]

    for cn in search_norms:
        for patterns, target in PHRASE_HINTS:
            base = target.split(".")[0]
            if base in CHI_EXCLUDED_TARGETS:
                continue
            if any_hint_match(cn, patterns):
                for entry in resolve_targets(target, cn, catalog, children):
                    if entry:
                        results[entry["ma"]] = entry

    return sorted(results.values(), key=lambda x: x["ma"])[:max_codes]


def map_chong_text_to_icds(text: str, catalog: dict, children: dict, max_codes: int = 15) -> list[dict]:
    """Map chống chỉ định: chỉ mã phản ánh lý do không được dùng thuốc."""
    results: dict[str, dict] = {}
    for clause in split_clauses(text):
        cn = norm(clause)
        if not cn:
            continue
        if RELATIVE_CONTRA_CLAUSE.search(cn):
            continue
        if not CHONG_CLAUSE_SIGNAL.search(cn):
            continue
        for patterns, target in CHONG_PHRASE_HINTS:
            if any_hint_match(cn, patterns):
                for entry in resolve_targets(target, cn, catalog, children):
                    if entry:
                        results[entry["ma"]] = entry
    return sorted(results.values(), key=lambda x: x["ma"])[:max_codes]


def remove_chong_overlap_with_chi(chi_entries: list[dict], chong_entries: list[dict]) -> list[dict]:
    chi_mas = {e["ma"] for e in chi_entries}
    return [e for e in chong_entries if e["ma"] not in chi_mas]


def apply_icd_fields_to_drug(d: dict, catalog: dict, children: dict) -> None:
    chi_entries = map_text_to_icds(d.get("chiDinh") or "", catalog, children)
    chong_entries = map_chong_text_to_icds(d.get("chongChiDinh") or "", catalog, children)
    chong_entries = remove_chong_overlap_with_chi(chi_entries, chong_entries)

    ma, ten, legacy = format_split_fields(chi_entries)
    d["maICDChiDinh"] = ma
    d["tenBenhICDChiDinh"] = ten
    d["goiYMaICDChiDinh"] = legacy

    ma, ten, legacy = format_split_fields(chong_entries)
    d["maICDChongChiDinh"] = ma
    d["tenBenhICDChongChiDinh"] = ten
    d["goiYMaICDChongChiDinh"] = legacy


def format_split_fields(entries: list[dict]) -> tuple[str, str, str]:
    if not entries:
        return "", "", ""
    ordered = sorted({e["ma"]: e for e in entries if e}.values(), key=lambda x: x["ma"])
    ma = ";".join(e["ma"] for e in ordered)
    ten = ";".join(e["ten"] for e in ordered)
    legacy = "; ".join(f"{e['ma']} — {e['ten']}" for e in ordered)
    return ma, ten, legacy


def parse_legacy_combined(text: str, catalog: dict) -> tuple[str, str]:
    mas: list[str] = []
    tens: list[str] = []
    for part in str(text or "").split(";"):
        part = part.strip()
        if not part:
            continue
        m = re.match(r"^([A-TV-ZU]\d{2}(?:\.\d{1,2})?\*?)\s*[—–\-]\s*(.+)$", part, re.I)
        if m:
            ma = m.group(1).upper()
            mas.append(ma)
            tens.append(m.group(2).strip())
            continue
        m2 = ICD_RE.match(part)
        if m2:
            ma = m2.group(1).upper()
            mas.append(ma)
            tens.append(catalog.get(ma, {}).get("ten", ""))
    return ";".join(mas), ";".join(t for t in tens if t)


def load_drugs_from_html(html_text: str) -> list[dict]:
    drugs: list[dict] = []
    for m in re.finditer(r'class="drugs-data-chunk"[^>]*>(\[.*?\])</script>', html_text, re.S):
        drugs.extend(json.loads(m.group(1)))
    return drugs


def patch_html_drug_chunks(html_text: str, drugs: list[dict]) -> str:
    chunks: list[list[dict]] = []
    for m in re.finditer(r'class="drugs-data-chunk"[^>]*>(\[.*?\])</script>', html_text, re.S):
        chunks.append(json.loads(m.group(1)))
    idx = 0
    new_chunks: list[list[dict]] = []
    for old in chunks:
        new_chunks.append(drugs[idx : idx + len(old)])
        idx += len(old)
    it = iter(new_chunks)

    def repl(m: re.Match) -> str:
        chunk = next(it)
        open_tag = m.group(0)[: m.group(0).index(">") + 1]
        return open_tag + json.dumps(chunk, ensure_ascii=False, separators=(",", ":")) + "</script>"

    new_html, n = re.subn(
        r'class="drugs-data-chunk"[^>]*>\[.*?\]</script>',
        repl,
        html_text,
        count=len(chunks),
        flags=re.S,
    )
    if n != len(chunks):
        raise SystemExit(f"Patch failed: {n}/{len(chunks)}")
    return new_html


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--xlsx", default=XLSX_DEFAULT)
    ap.add_argument("--dry-run", action="store_true")
    args = ap.parse_args()

    html_path = next(WORKSPACE.glob(HTML_GLOB))
    html_text = html_path.read_text(encoding="utf-8")
    loaded = try_load_xlsx(args.xlsx)
    if loaded:
        catalog, entries, children = loaded
        print("ICD catalog: Excel", len(entries), "ma")
    else:
        catalog, entries, children = load_icd_from_html(html_text)
        print("ICD catalog: embedded HTML", len(entries), "ma")

    drugs = load_drugs_from_html(html_text)
    print("Thuoc:", len(drugs))

    stats = {"chi": 0, "chong": 0, "det_chi": 0, "det_chong": 0, "multi_e11": 0, "overlap": 0}

    for d in drugs:
        apply_icd_fields_to_drug(d, catalog, children)

        chi_set = set((d.get("maICDChiDinh") or "").split(";")) - {""}
        chong_set = set((d.get("maICDChongChiDinh") or "").split(";")) - {""}
        if chi_set & chong_set:
            stats["overlap"] += 1

        if d.get("maICDChiDinh"):
            stats["chi"] += 1
            if "E11.8" in d["maICDChiDinh"] and "E11.9" in d["maICDChiDinh"]:
                stats["multi_e11"] += 1
            if any("." in p for p in d["maICDChiDinh"].split(";")):
                stats["det_chi"] += 1
        if d.get("maICDChongChiDinh"):
            stats["chong"] += 1
            if any("." in p for p in d["maICDChongChiDinh"].split(";")):
                stats["det_chong"] += 1

    print("Stats:", stats)
    for name in ("Losartan", "Metformin", "Amoxicilin", "Irbesartan", "Abacavir"):
        drug = next((x for x in drugs if name in (x.get("tenHoatChat") or "")), None)
        if not drug:
            continue
        print(f"\n{name}:")
        print("  ma CHI:", drug.get("maICDChiDinh", "")[:200])
        print("  ten CHI:", drug.get("tenBenhICDChiDinh", "")[:200])
        print("  ma CHONG:", drug.get("maICDChongChiDinh", "")[:160])

    if args.dry_run:
        print("\nDry-run — khong ghi file.")
        return

    new_html = patch_html_drug_chunks(html_text, drugs)
    if "</html>" not in new_html:
        raise SystemExit("HTML truncated!")
    html_path.write_text(new_html, encoding="utf-8")
    print("\nDa cap nhat du lieu thuoc:", html_path.name)


if __name__ == "__main__":
    main()
