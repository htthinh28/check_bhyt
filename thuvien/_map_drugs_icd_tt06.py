# -*- coding: utf-8 -*-
"""
Map goiYMaICDChiDinh / goiYMaICDChongChiDinh từ chiDinh/chongChiDinh
+ danh mục ICD-10 TT06 (Excel). Ưu tiên mã chi tiết (4–5 ký tự).

Chạy: python _map_drugs_icd_tt06.py [--dry-run]
"""
from __future__ import annotations

import argparse
import json
import re
import sys
import unicodedata
from collections import defaultdict
from pathlib import Path

try:
    import openpyxl
except ImportError:
    print("Can cai: pip install openpyxl", file=sys.stderr)
    sys.exit(1)

sys.stdout.reconfigure(encoding="utf-8")

XLSX_DEFAULT = r"c:\Users\admin\Downloads\Phu luc Bang danh muc ICD10_FINAL  (1).xlsx"
THU_VIEN = Path(r"C:\Users\admin\Desktop\Thu vien")

# (patterns trong text đã chuẩn hóa, tiền tố ICD 3 ký tự hoặc mã đích)
PHRASE_HINTS: list[tuple[list[str], str]] = [
    (["nhiem hiv", " hiv", "aids", "suy giam mien dich"], "B24"),
    (["viem tai giua", "otitis media"], "H66"),
    (["viem xoang", "sinusitis"], "J01"),
    (["viem amidan", "viem amydan", "viem hong", "viem amiđan", "viem amidan"], "J03"),
    (["viem phoi", "pneumonia", "viem phoi mac phai", "viem phoi benh vien"], "J18"),
    (["viem phe quan man", "dot bung phat", "copd"], "J44"),
    (["nhiem trung da", "nhiem khuan da", "viem da va mo", "ap xe rang", "viem mo te bao lan toa"], "L03"),
    (["viem bang quang"], "N30"),
    (["viem than", "be than", "bể thận", "viem than-be than"], "N10"),
    (["loet da day", "loet da day-ta trang", "helicobacter", "hpylori"], "K25"),
    (["benh lyme", "lyme"], "A69.2"),
    (["anthrax", "bacillus anthracis", "benh than do vi khuan", "phoi nhiem than"], "A22"),
    (["benh lau", "gonorrh", "gonococc"], "A54"),
    (["viem mang nao", "meningitis"], "G03"),
    (["tang huyet ap", "tăng huyết áp", "tăng huyết áp nguyên phát"], "I10"),
    (["suy tim", "suy tim mat bu", "suy tim sung huyet"], "I50"),
    (["nhoi mau co tim", "nhiem mau co tim", "st chênh lên", "st chech len"], "I21"),
    (["dot quy", "thieu mau cuc bo", "cơn thiếu máu cục bộ"], "I63"),
    (["dai thao duong typ 2", "dtd typ 2", "dtd type 2", "diabetes typ 2", "đái tháo đường typ 2"], "E11"),
    (["ung thu vu", "ung thư vú", "carcinoma vu"], "C50"),
    (["ung thu phoi", "ung thư phổi", "carcinoma phoi"], "C34"),
    (["ung thu tuyen tien liet", "ung thư tuyến tiền liệt", "prostate"], "C61"),
    (["ung thu da day", "ung thư dạ dày", "adenocarcinoma da day"], "C16"),
    (["ung thu dau", "ung thư đầu", "ung thu co", "đầu - cổ"], "C00"),
    (["chung dau", "chứng đau", "giam dau", "giảm đau", " dau nhe", " đau vua"], "R52"),
    ([" ha sot", "hạ sốt", " sot "], "R50"),
    (["suy gan", "suy gan nang", "suy gan vua"], "K72"),
    (["suy than", "suy thận", "gfr", "loc cau than"], "N18"),
    (["mang thai", "phu nu mang thai", "thai ky", "co thai", "thai kỳ"], "Z33"),
    (["qua man", "mẫn cảm", "di ung", "dị ứng"], "T78"),
    ([" lao", "lao:", "tuberculosis", "mycobacterium tuberculosis"], "A15"),
    (["dich hai", "peste", "yersinia pestis"], "A20"),
    (["brucella"], "A23"),
    (["sot vang", "yellow fever"], "A95"),
    (["viem khop", "nhiem khuan xuong", "nhiễm khuẩn xương"], "M00"),
    (["viem mang bung", "viêm màng bụng"], "K65"),
    (["nhiem khuan huyet", "sepsis", "sot giam bach cau", "nhiễm khuẩn huyết"], "A41"),
    (["hen ", " hen", "asthma", "hen phe quan"], "J45"),
    (["viem tai ngoai", "viêm tai ngoài"], "H60"),
    (["viem tai giua mu", "viêm tai giữa mủ"], "H66"),
    (["viem phe quan cap", "viêm phế quản cấp"], "J20"),
    (["nhiem khuan duong tieu", "viem bang quang cap", "viêm bàng quang cấp"], "N30"),
    (["viem than-be than cap", "viem than cap", "viêm thận-bể thận"], "N10"),
    (["viem phoi benh vien", "viêm phổi bệnh viện"], "J15"),
    (["viem phoi xo hoa", "xơ hóa"], "J84"),
    (["can thiep mach vanh", "mạch vành qua da", "pci"], "I25"),
    (["dong mach ngoai vi", "bệnh động mạch ngoại vi"], "I73"),
    (["benh than o", "benh than dai thao duong", "bệnh thận đái tháo đường", "protein nieu"], "N08"),
    (["co giat", "co giật", "epilep"], "G40"),
    (["dau than kinh", "đau thần kinh", "neuropathic"], "G62"),
    (["benh parkinson", "parkinson"], "G20"),
    (["rung nhi", "rung nhĩ", "fibrillation"], "I48"),
    (["viem loet dai trang", "colitis", "viêm loét đại tràng"], "K51"),
    (["viem da day", "viêm dạ dày"], "K29"),
    (["tang lipid", "tăng lipid", "cholesterol", "rối loạn lipid"], "E78"),
    (["basedow", "cường giáp", "cuong giap"], "E05"),
    (["suy giap", "suy giáp"], "E03"),
    (["viem gan b", "viêm gan b", "hepatitis b"], "B18"),
    (["viem gan c", "viêm gan c", "hepatitis c"], "B18.2"),
    (["benh giang mai", "giang mai", "syphilis"], "A53"),
    (["nong do benzodiazepin", "ngộ độc benzodiazepin", "ngu doc benzodiazepin"], "T42"),
    (["ngu doc", "ngộ độc", "qua lieu"], "T50"),
    (["teo day than kinh thi giac leber"], "H47"),
    (["giam thi luc do thuoc la"], "H53"),
    (["suy tim cung luong cao", "soc nhiem khuan"], "R57"),
    (["hep eo dong mach chu", "hẹp eo động mạch chủ"], "Q25"),
    (["benh da hong cau", "đa hồng cầu"], "D45"),
    (["bach cau tuy bao man", "bạch cầu tủy bào mạn"], "C92"),
    (["xơ hóa tủy", "xo hoa tuy"], "D75"),
    (["tang tieu cau", "tăng tiểu cầu"], "D47"),
    (["benh te bao goc", "tế bào gốc"], "Z51"),
    (["viem da day do thuoc", "viêm dạ dày do thuốc"], "K29"),
    (["viem khop dang thap", "viêm khớp dạng thấp"], "M05"),
    (["thap khop", "thấp khớp"], "M06"),
    (["gout", "gút"], "M10"),
    (["loang xuong", "loãng xương"], "M81"),
    (["benh nhiem trung duong sinh duc", "herpes simplex"], "A60"),
    (["viem gan man", "viêm gan mạn"], "K73"),
    (["viem phoi do vi khuan", "viêm phổi do vi khuẩn"], "J15"),
    (["viem phoi do virus", "viêm phổi do virus"], "J12"),
    (["viem phoi do vi sinh vat ky sinh"], "J16"),
    (["viem phoi do vi khuan khac"], "J18"),
    (["viem phoi do vi khuan khac va khong xac dinh"], "J18"),
    (["viem phoi do vi khuan khac va khong xac dinh"], "J18"),
]

SKIP_CLAUSE = re.compile(
    r"^(cach dung|lieu dung|nguoi lon|tre em|phu nu|can |khong tu|phoi hop voi thuoc|"
    r"duoc chi dinh|duoc su dung|chi dinh|chong chi dinh|muc do|liều|uống|tiêm|truyền|"
    r"phac do|phác đồ|co che|co che|tiêu chuẩn|lợi ích|người bệnh có thể)\b",
    re.I,
)

ICD_RE = re.compile(r"^([A-TV-ZU]\d{2}(?:\.\d{1,2})?)")


def norm(s: str) -> str:
    s = unicodedata.normalize("NFD", str(s or ""))
    s = "".join(c for c in s if unicodedata.category(c) != "Mn")
    s = s.replace("đ", "d").replace("Đ", "D")
    s = re.sub(r"[^\w\s.]", " ", s.lower())
    s = re.sub(r"\s+", " ", s).strip()
    return s


def split_clauses(text: str) -> list[str]:
    if not text:
        return []
    t = str(text).replace("\r\n", "\n")
    parts: list[str] = []
    for line in re.split(r"[\n;]+", t):
        line = re.sub(r"^[•\-\*·▪–—]\s*", "", line.strip())
        line = re.sub(r"^\d+[.)]\s*", "", line)
        if line:
            parts.append(line)
    return parts


def load_icd_catalog(xlsx_path: str):
    wb = openpyxl.load_workbook(xlsx_path, read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]
    catalog: dict[str, dict] = {}
    entries: list[dict] = []
    for row in ws.iter_rows(min_row=5, values_only=True):
        ma = str(row[17] or "").strip().upper()
        if not ma or not ICD_RE.match(ma):
            continue
        ten = str(row[21] or "").strip()
        flag4 = bool(str(row[25] or "").strip())
        e = {"ma": ma, "ten": ten, "norm_ten": norm(ten), "flag4": flag4}
        catalog[ma] = e
        entries.append(e)

    children: dict[str, list[dict]] = defaultdict(list)
    for e in entries:
        children[e["ma"].split(".")[0]].append(e)

    return catalog, entries, children


def score_match(clause_norm: str, entry: dict) -> float:
    ten = entry["norm_ten"]
    if not ten or len(ten) < 6:
        return 0.0
    if ten in clause_norm:
        return 200.0 + len(ten)
    # khớp cụm ≥ 12 ký tự liên tiếp trong tên bệnh
    words = [w for w in ten.split() if len(w) >= 5]
    hits = sum(1 for w in words if w in clause_norm)
    if hits >= 2:
        return 90.0 + hits * 8
    return 0.0


def pick_best_under_prefix(prefix: str, clause_norm: str, catalog: dict, children: dict) -> dict | None:
    prefix = prefix.upper()
    if prefix in catalog and "." in prefix:
        return catalog[prefix]
    base = prefix.split(".")[0]
    cands = children.get(base, [])
    dotted = [c for c in cands if "." in c["ma"]]
    if not dotted:
        return catalog.get(prefix) or catalog.get(base)
    best = None
    best_sc = 0.0
    for c in dotted:
        sc = score_match(clause_norm, c)
        if sc > best_sc:
            best_sc = sc
            best = c
    if best and best_sc >= 35:
        return best
    for c in dotted:
        if c["ma"].endswith(".9"):
            return c
    return dotted[0] if dotted else catalog.get(base)


def resolve_detailed(entry: dict | None, clause_norm: str, catalog: dict, children: dict) -> dict | None:
    if not entry:
        return None
    ma = entry["ma"]
    base = ma.split(".")[0]
    dotted_children = [c for c in children.get(base, []) if "." in c["ma"]]
    if "." not in ma and (entry.get("flag4") or len(dotted_children) > 1):
        picked = pick_best_under_prefix(base, clause_norm, catalog, children)
        return picked or entry
    return entry


def parse_codes_from_field(text: str) -> list[str]:
    out: list[str] = []
    for part in str(text or "").split(";"):
        m = ICD_RE.match(part.strip())
        if m:
            out.append(m.group(1).upper())
    return out


def format_icd_field(entries: list[dict]) -> str:
    if not entries:
        return ""
    uniq: dict[str, dict] = {}
    for e in entries:
        if e:
            uniq[e["ma"]] = e
    ordered = sorted(uniq.values(), key=lambda x: x["ma"])
    return "; ".join(f"{e['ma']} — {e['ten']}" for e in ordered)


def map_text_to_icds(text: str, existing_field: str, catalog, children, max_codes=20) -> str:
    results: dict[str, dict] = {}
    full_norm = norm(text)

    # Loại câu phủ định (vd. "không có tác dụng trị thấp khớp")
    if re.search(r"khong (co|duoc) (tac dung|dung|chi dinh)", full_norm):
        work_norm = full_norm
        for seg in re.split(r"[.;:\n]", full_norm):
            sn = seg.strip()
            if re.search(r"khong (co|duoc) (tac dung|dung|chi dinh)", sn):
                work_norm = work_norm.replace(sn, " ")
    else:
        work_norm = full_norm

    for patterns, target in PHRASE_HINTS:
        if any(p in work_norm for p in patterns):
            hit = pick_best_under_prefix(target, work_norm, catalog, children)
            if hit:
                results[hit["ma"]] = hit

    return format_icd_field(list(results.values())[:max_codes])


def load_drugs_from_html(html_text: str) -> list[dict]:
    drugs: list[dict] = []
    for m in re.finditer(r'class="drugs-data-chunk"[^>]*>(\[.*?\])</script>', html_text, re.S):
        drugs.extend(json.loads(m.group(1)))
    return drugs


def patch_html_drug_chunks(html_text: str, drugs: list[dict]) -> str:
    chunks: list[list[dict]] = []
    for m in re.finditer(r'class="drugs-data-chunk"[^>]*>(\[.*?\])</script>', html_text, re.S):
        chunks.append(json.loads(m.group(1)))
    idx = 0
    new_chunks: list[list[dict]] = []
    for old in chunks:
        new_chunks.append(drugs[idx : idx + len(old)])
        idx += len(old)
    it = iter(new_chunks)

    def repl(m: re.Match) -> str:
        chunk = next(it)
        open_tag = m.group(0)[: m.group(0).index(">") + 1]
        return open_tag + json.dumps(chunk, ensure_ascii=False, separators=(",", ":")) + "</script>"

    new_html, n = re.subn(
        r'class="drugs-data-chunk"[^>]*>\[.*?\]</script>',
        repl,
        html_text,
        count=len(chunks),
        flags=re.S,
    )
    if n != len(chunks):
        raise SystemExit(f"Patch failed: {n}/{len(chunks)}")
    return new_html


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--xlsx", default=XLSX_DEFAULT)
    ap.add_argument("--dry-run", action="store_true")
    args = ap.parse_args()

    html_path = next(THU_VIEN.glob("*(1)*.html"))
    catalog, entries, children = load_icd_catalog(args.xlsx)
    html_text = html_path.read_text(encoding="utf-8")
    drugs = load_drugs_from_html(html_text)
    print("ICD:", len(entries), "ma | Thuoc:", len(drugs))

    stats = {"chi": 0, "chong": 0, "det_chi": 0, "det_chong": 0}

    for i, d in enumerate(drugs):
        old_chi = d.get("goiYMaICDChiDinh") or ""
        old_chong = d.get("goiYMaICDChongChiDinh") or ""
        chi = map_text_to_icds(d.get("chiDinh") or "", old_chi, catalog, children)
        chong = map_text_to_icds(d.get("chongChiDinh") or "", old_chong, catalog, children)
        d["goiYMaICDChiDinh"] = chi
        d["goiYMaICDChongChiDinh"] = chong
        if chi:
            stats["chi"] += 1
            if any("." in p.split("—")[0].strip() for p in chi.split(";")):
                stats["det_chi"] += 1
        if chong:
            stats["chong"] += 1
            if any("." in p.split("—")[0].strip() for p in chong.split(";")):
                stats["det_chong"] += 1

    print("Stats:", stats)
    for name in ("Amoxicilin", "Docetaxel", "Losartan", "Abacavir", "Paracetamol"):
        d = next((x for x in drugs if name in (x.get("tenHoatChat") or "")), None)
        if d:
            print(f"\n{name}:")
            print("  CHI:", (d.get("goiYMaICDChiDinh") or "")[:350])
            print("  CHONG:", (d.get("goiYMaICDChongChiDinh") or "")[:200])

    if args.dry_run:
        print("\nDry-run — khong ghi file.")
        return

    new_html = patch_html_drug_chunks(html_text, drugs)
    if not new_html.rstrip().endswith("</html>"):
        raise SystemExit("HTML truncated!")
    html_path.write_text(new_html, encoding="utf-8")
    print("\nDa cap nhat:", html_path.name)


if __name__ == "__main__":
    main()
