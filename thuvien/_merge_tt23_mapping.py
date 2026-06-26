# -*- coding: utf-8 -*-
"""Map TT23 Excel vào danh mục DVKT, gắn thẻ nguồn văn bản."""
from __future__ import annotations

import json
import re
import unicodedata
from pathlib import Path

import openpyxl

HTML_PATH = Path(r"G:\My Drive\Thu vien (1)\dich vụ kỹ thuật.html")
XLSX_PATH = Path(r"d:\BHYT\Danh muc TT23\Danh muc theo thong tu 23.xlsx")

LABEL_QD7603 = "Quyết định số 7603/QĐ-BYT"
LABEL_TT23 = "Thông tư số 23/2024/TT-BYT"
LABEL_TT23_PL1 = f"{LABEL_TT23} — Phụ lục 1 (Danh mục kỹ thuật)"
LABEL_TT23_PL2 = f"{LABEL_TT23} — Phụ lục 2 (Danh mục phẫu thuật)"


def norm_text(value) -> str:
    if value is None:
        return ""
    text = str(value).strip().lower()
    text = unicodedata.normalize("NFD", text)
    text = "".join(c for c in text if unicodedata.category(c) != "Mn")
    text = re.sub(r"\s+", " ", text)
    return text


def norm_code(value) -> str:
    if value is None or value == "":
        return ""
    try:
        num = float(str(value).replace(",", "."))
        text = f"{num:.10f}".rstrip("0").rstrip(".")
        return text
    except (TypeError, ValueError):
        return str(value).strip()


def name_similarity(a: str, b: str) -> float:
    na, nb = norm_text(a), norm_text(b)
    if not na or not nb:
        return 0.0
    if na == nb:
        return 1.0
    if na in nb or nb in na:
        return 0.88
    ta, tb = set(na.split()), set(nb.split())
    if not ta or not tb:
        return 0.0
    return len(ta & tb) / len(ta | tb)


def pick_best(candidates: list[dict], ten_ref: str) -> tuple[dict, str]:
    if len(candidates) == 1:
        score = name_similarity(ten_ref, candidates[0]["ten"])
        if score >= 0.85:
            return candidates[0], "Cao"
        if score >= 0.55:
            return candidates[0], "Trung bình"
        return candidates[0], "Thấp"
    scored = [(name_similarity(ten_ref, c["ten"]), c) for c in candidates]
    scored.sort(key=lambda x: x[0], reverse=True)
    best_score, best = scored[0]
    if best_score >= 0.9:
        return best, "Cao"
    if best_score >= 0.6:
        return best, "Trung bình"
    return best, "Thấp"


def pick_best_7603(candidates: list[dict], ten_ref: str) -> tuple[dict, str]:
    wrapped = [{**c, "ten": c.get("tenTT43", "")} for c in candidates]
    return pick_best(wrapped, ten_ref)


def mapping_note(method: str, ma43: str, ten: str, match: dict, pl_label: str) -> str:
    parts = [f"Liên kết {pl_label} theo {method}"]
    if ma43:
        parts.append(f"mã TT43/TT23: {ma43}")
    parts.append(f"tên QĐ7603: «{ten}»")
    parts.append(f"tên TT23: «{match['ten']}»")
    return " · ".join(parts)


def match_to_tt23(
    ma43: str,
    ten: str,
    by_ma_pl1: dict,
    by_ten_pl1: dict,
    by_ma_pl2: dict,
    by_ten_pl2: dict,
) -> tuple[dict | None, str, str, str]:
    ten_n = norm_text(ten)
    code = norm_code(ma43)

    if code and code in by_ma_pl1:
        match, conf = pick_best(by_ma_pl1[code], ten)
        if name_similarity(ten, match["ten"]) >= 0.45:
            method = "mã + tên" if name_similarity(ten, match["ten"]) >= 0.85 else "mã (ưu tiên tên gần đúng)"
            return match, "TT23_PL1", method, conf

    if ten_n and ten_n in by_ten_pl1:
        cands = by_ten_pl1[ten_n]
        match = cands[0]
        method = "tên chính xác" if len(cands) == 1 else "tên (nhiều mã trùng tên)"
        return match, "TT23_PL1", method, "Cao" if len(cands) == 1 else "Trung bình"

    if code and code in by_ma_pl2:
        match, conf = pick_best(by_ma_pl2[code], ten)
        if name_similarity(ten, match["ten"]) >= 0.45:
            method = "mã + tên" if name_similarity(ten, match["ten"]) >= 0.85 else "mã PL2 (ưu tiên tên gần đúng)"
            return match, "TT23_PL2", method, conf

    for cands in by_ten_pl2.values():
        for c in cands:
            if name_similarity(ten, c["ten"]) >= 0.92:
                return c, "TT23_PL2", "tên gần đúng PL2", "Trung bình"

    return None, "", "", "Không khớp"


def read_tt23() -> tuple[list[dict], list[dict]]:
    wb = openpyxl.load_workbook(XLSX_PATH, read_only=True, data_only=True)
    pl1, pl2 = [], []
    for row in wb["Phu luc 1"].iter_rows(min_row=2, values_only=True):
        if row[1] in (None, ""):
            continue
        stt = int(row[0]) if row[0] not in (None, "") else len(pl1) + 1
        ma = norm_code(row[1])
        pl1.append(
            {
                "stt": stt,
                "maKyThuat": ma,
                "chuong": str(row[2] or "").strip(),
                "tenKyThuat": str(row[3] or "").strip(),
                "phuLuc": str(row[4] or "PL01").strip(),
                "_rowId": f"tt23pl1-{ma}-{stt}",
            }
        )
    for row in wb["Phu luc 2"].iter_rows(min_row=2, values_only=True):
        if row[3] in (None, ""):
            continue
        stt = int(row[0]) if row[0] not in (None, "") else len(pl2) + 1
        ma = norm_code(row[3])
        pl2.append(
            {
                "stt": stt,
                "sttChuong": int(row[1]) if row[1] not in (None, "") else "",
                "chuong": str(row[2] or "").strip(),
                "maKyThuat": ma,
                "tenKyThuat": str(row[4] or "").strip(),
                "phuLuc": str(row[5] or "PK02").strip(),
                "_rowId": f"tt23pl2-{ma}-{stt}",
            }
        )
    wb.close()
    return pl1, pl2


def index_tt23(pl1: list[dict], pl2: list[dict]):
    by_ma_pl1, by_ten_pl1 = {}, {}
    by_ma_pl2, by_ten_pl2 = {}, {}
    for r in pl1:
        entry = {"maKyThuat": r["maKyThuat"], "ten": r["tenKyThuat"], "chuong": r["chuong"], "phuLuc": r["phuLuc"], "_rowId": r["_rowId"]}
        by_ma_pl1.setdefault(r["maKyThuat"], []).append(entry)
        by_ten_pl1.setdefault(norm_text(r["tenKyThuat"]), []).append(entry)
    for r in pl2:
        entry = {"maKyThuat": r["maKyThuat"], "ten": r["tenKyThuat"], "chuong": r["chuong"], "phuLuc": r["phuLuc"], "_rowId": r["_rowId"]}
        by_ma_pl2.setdefault(r["maKyThuat"], []).append(entry)
        by_ten_pl2.setdefault(norm_text(r["tenKyThuat"]), []).append(entry)
    return by_ma_pl1, by_ten_pl1, by_ma_pl2, by_ten_pl2


def load_pl1_from_html(html: str) -> list[dict]:
    m = re.search(
        r"const INITIAL_PL1_DATA = (\[.*?\]);\s*\n\n\s*const INITIAL_PL2_COLUMNS",
        html,
        re.S,
    )
    if not m:
        raise SystemExit("Khong tim thay INITIAL_PL1_DATA")
    return json.loads(m.group(1))


def enrich_pl1(
    pl1_7603: list[dict],
    by_ma_pl1,
    by_ten_pl1,
    by_ma_pl2,
    by_ten_pl2,
) -> tuple[list[dict], dict]:
    stats = {"Cao": 0, "Trung bình": 0, "Thấp": 0, "Không khớp": 0, "TT23_PL1": 0, "TT23_PL2": 0}
    reverse: dict[str, list[dict]] = {}

    for item in pl1_7603:
        ma43 = item.get("maTT43", "")
        ten = item.get("tenTT43", "")
        match, link_tag, method, conf = match_to_tt23(
            ma43, ten, by_ma_pl1, by_ten_pl1, by_ma_pl2, by_ten_pl2
        )

        item["theNguonGoc"] = "QD_7603"
        item["vanBanNguon"] = LABEL_QD7603

        if match:
            pl_label = LABEL_TT23_PL1 if link_tag == "TT23_PL1" else LABEL_TT23_PL2
            item["theLienKetTT23"] = link_tag
            item["maKyThuatTT23"] = match["maKyThuat"]
            item["tenKyThuatTT23"] = match["ten"]
            item["chuongTT23"] = match["chuong"]
            item["phuLucTT23"] = match["phuLuc"]
            item["doTinCayMapping"] = conf
            item["ghiChuMapping"] = mapping_note(method, norm_code(ma43), ten, match, pl_label)
            stats[conf] += 1
            stats[link_tag] += 1
            reverse.setdefault(match["_rowId"], []).append(item)
        else:
            item["theLienKetTT23"] = ""
            item["maKyThuatTT23"] = ""
            item["tenKyThuatTT23"] = ""
            item["chuongTT23"] = ""
            item["phuLucTT23"] = ""
            item["doTinCayMapping"] = "Không khớp"
            item["ghiChuMapping"] = (
                f"Chỉ có trong {LABEL_QD7603}; chưa tìm thấy mã/tên tương ứng trong {LABEL_TT23}"
            )
            stats["Không khớp"] += 1

    return pl1_7603, stats, reverse


def build_tt23_datasets(
    pl1_tt23: list[dict],
    pl2_tt23: list[dict],
    reverse: dict[str, list[dict]],
    by_ma_pl1,
    by_ten_pl1,
    by_ma_pl2,
    by_ten_pl2,
    pl1_7603: list[dict],
) -> tuple[list[dict], list[dict]]:
    idx_ma_7603, idx_ten_7603 = {}, {}
    for r in pl1_7603:
        c = norm_code(r.get("maTT43", ""))
        if c:
            idx_ma_7603.setdefault(c, []).append(r)
        idx_ten_7603.setdefault(norm_text(r.get("tenTT43", "")), []).append(r)

    out1 = []
    for r in pl1_tt23:
        row = dict(r)
        row["theNguonGoc"] = "TT23_PL1"
        row["vanBanNguon"] = LABEL_TT23_PL1
        links = reverse.get(r["_rowId"], [])
        if links:
            best = links[0]
            row["lienKetQD7603"] = best["maTuongDuong"]
            row["maTuongDuongQD7603"] = best["maTuongDuong"]
            row["tenTT43QD7603"] = best["tenTT43"]
            row["doTinCayMapping"] = best["doTinCayMapping"]
            row["ghiChuMapping"] = f"Đã liên kết ngược với {LABEL_QD7603}: {best['maTuongDuong']}"
        else:
            code = r["maKyThuat"]
            ten = r["tenKyThuat"]
            hit = None
            if code in idx_ma_7603:
                hit, conf = pick_best_7603(idx_ma_7603[code], ten)
            elif norm_text(ten) in idx_ten_7603:
                hit = idx_ten_7603[norm_text(ten)][0]
                conf = "Trung bình"
            if hit:
                row["lienKetQD7603"] = hit["maTuongDuong"]
                row["maTuongDuongQD7603"] = hit["maTuongDuong"]
                row["tenTT43QD7603"] = hit["tenTT43"]
                row["doTinCayMapping"] = conf if hit else "Thấp"
                row["ghiChuMapping"] = f"Liên kết ngược theo mã/tên với {LABEL_QD7603}"
            else:
                row["lienKetQD7603"] = ""
                row["maTuongDuongQD7603"] = ""
                row["tenTT43QD7603"] = ""
                row["doTinCayMapping"] = "Chỉ TT23"
                row["ghiChuMapping"] = f"Chỉ có trong {LABEL_TT23_PL1}, không có trong {LABEL_QD7603}"
        out1.append(row)

    out2 = []
    for r in pl2_tt23:
        row = dict(r)
        row["theNguonGoc"] = "TT23_PL2"
        row["vanBanNguon"] = LABEL_TT23_PL2
        code = r["maKyThuat"]
        ten = r["tenKyThuat"]
        hit = None
        conf = "Chỉ TT23"
        if code in idx_ma_7603:
            hit, conf = pick_best_7603(idx_ma_7603[code], ten)
        elif norm_text(ten) in idx_ten_7603:
            hit = idx_ten_7603[norm_text(ten)][0]
            conf = "Trung bình"
        if hit:
            row["lienKetQD7603"] = hit["maTuongDuong"]
            row["maTuongDuongQD7603"] = hit["maTuongDuong"]
            row["tenTT43QD7603"] = hit["tenTT43"]
            row["doTinCayMapping"] = conf
            row["ghiChuMapping"] = f"Liên kết với {LABEL_QD7603} theo mã/tên"
        else:
            row["lienKetQD7603"] = ""
            row["maTuongDuongQD7603"] = ""
            row["tenTT43QD7603"] = ""
            row["doTinCayMapping"] = "Chỉ TT23"
            row["ghiChuMapping"] = f"Chỉ có trong {LABEL_TT23_PL2}, không có trong {LABEL_QD7603}"
        out2.append(row)

    return out1, out2


def patch_block(html: str, name: str, data, next_marker: str) -> str:
    js = json.dumps(data, ensure_ascii=False, indent=2)
    pat = rf"const {name} = \[.*?\];\s*(?=\s*{re.escape(next_marker)})"
    repl = f"const {name} = {js};\n\n    "
    new_html, n = re.subn(pat, repl, html, count=1, flags=re.S)
    if n != 1:
        raise SystemExit(f"Khong patch duoc {name} (matches={n})")
    return new_html


def insert_after_marker(html: str, marker: str, insertion: str) -> str:
    if insertion.strip() in html:
        return html
    idx = html.find(marker)
    if idx == -1:
        raise SystemExit(f"Khong tim thay marker: {marker}")
    return html[:idx] + insertion + html[idx:]


def main():
    if not XLSX_PATH.is_file():
        raise SystemExit(f"Khong tim thay: {XLSX_PATH}")
    html = HTML_PATH.read_text(encoding="utf-8")

    pl1_tt23, pl2_tt23 = read_tt23()
    by_ma_pl1, by_ten_pl1, by_ma_pl2, by_ten_pl2 = index_tt23(pl1_tt23, pl2_tt23)
    pl1_7603 = load_pl1_from_html(html)
    pl1_7603, stats, reverse = enrich_pl1(
        pl1_7603, by_ma_pl1, by_ten_pl1, by_ma_pl2, by_ten_pl2
    )
    tt23_pl1, tt23_pl2 = build_tt23_datasets(
        pl1_tt23, pl2_tt23, reverse, by_ma_pl1, by_ten_pl1, by_ma_pl2, by_ten_pl2, pl1_7603
    )

    print(f"PL1 7603 enriched: {len(pl1_7603)}")
    print("Mapping stats:", stats)
    print(f"TT23 PL1: {len(tt23_pl1)} | TT23 PL2: {len(tt23_pl2)}")
    linked1 = sum(1 for r in tt23_pl1 if r.get("lienKetQD7603"))
    linked2 = sum(1 for r in tt23_pl2 if r.get("lienKetQD7603"))
    print(f"TT23 linked to QD7603: PL1={linked1} PL2={linked2}")

    html = patch_block(html, "INITIAL_PL1_DATA", pl1_7603, "const INITIAL_PL2_COLUMNS")

    tt23_cols_pl1 = [
        {"key": "stt", "label": "STT", "type": "number", "required": True},
        {"key": "maKyThuat", "label": "Mã kỹ thuật TT23", "type": "text", "required": True},
        {"key": "tenKyThuat", "label": "Tên kỹ thuật TT23", "type": "text", "required": True},
        {"key": "chuong", "label": "Chương/Chuyên khoa", "type": "text"},
        {"key": "theNguonGoc", "label": "Thẻ nguồn", "type": "text"},
        {"key": "vanBanNguon", "label": "Văn bản gốc", "type": "text"},
        {"key": "lienKetQD7603", "label": "Mã QĐ7603 liên kết", "type": "text"},
        {"key": "tenTT43QD7603", "label": "Tên QĐ7603 liên kết", "type": "text"},
        {"key": "doTinCayMapping", "label": "Độ tin cậy mapping", "type": "text"},
        {"key": "ghiChuMapping", "label": "Ghi chú mapping", "type": "text"},
    ]
    tt23_cols_pl2 = [
        {"key": "stt", "label": "STT PL2", "type": "number", "required": True},
        {"key": "maKyThuat", "label": "Mã kỹ thuật TT23", "type": "text", "required": True},
        {"key": "tenKyThuat", "label": "Tên phẫu thuật TT23", "type": "text", "required": True},
        {"key": "chuong", "label": "Chương", "type": "text"},
        {"key": "theNguonGoc", "label": "Thẻ nguồn", "type": "text"},
        {"key": "vanBanNguon", "label": "Văn bản gốc", "type": "text"},
        {"key": "lienKetQD7603", "label": "Mã QĐ7603 liên kết", "type": "text"},
        {"key": "tenTT43QD7603", "label": "Tên QĐ7603 liên kết", "type": "text"},
        {"key": "doTinCayMapping", "label": "Độ tin cậy mapping", "type": "text"},
        {"key": "ghiChuMapping", "label": "Ghi chú mapping", "type": "text"},
    ]

    tt23_bundle = (
        f"    const INITIAL_TT23_PL1_COLUMNS = {json.dumps(tt23_cols_pl1, ensure_ascii=False, indent=2)};\n\n"
        f"    const INITIAL_TT23_PL1_DATA = {json.dumps(tt23_pl1, ensure_ascii=False, indent=2)};\n\n"
        f"    const INITIAL_TT23_PL2_COLUMNS = {json.dumps(tt23_cols_pl2, ensure_ascii=False, indent=2)};\n\n"
        f"    const INITIAL_TT23_PL2_DATA = {json.dumps(tt23_pl2, ensure_ascii=False, indent=2)};\n\n"
    )
    html = insert_after_marker(html, "    // === HỆ THỐNG TRẠNG THÁI TOÀN CỤC (STATE) ===", tt23_bundle)

    HTML_PATH.write_text(html, encoding="utf-8")
    print(f"Da cap nhat HTML ({len(html) // 1024} KB)")


if __name__ == "__main__":
    main()
