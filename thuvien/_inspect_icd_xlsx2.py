# -*- coding: utf-8 -*-
import sys
sys.stdout.reconfigure(encoding='utf-8')
import openpyxl

p = r"c:\Users\admin\Downloads\Phu luc Bang danh muc ICD10_FINAL  (1).xlsx"
wb = openpyxl.load_workbook(p, read_only=True, data_only=True)
ws = wb[wb.sheetnames[0]]
rows = list(ws.iter_rows(min_row=4, max_row=20, values_only=True))
for i, r in enumerate(rows, start=4):
    ma = str(r[17] or '').strip() if len(r) > 17 else ''
    ten = str(r[21] or '') if len(r) > 21 else ''
    flag4 = r[25] if len(r) > 25 else None
    print(i, ma, '|', ten[:60], '| flag4=', flag4)

# stats
total = 0
with_ma = 0
with_dot = 0
parent_only = 0
from collections import defaultdict
parents = defaultdict(list)
for row in ws.iter_rows(min_row=4, values_only=True):
    total += 1
    ma = str(row[17] or '').strip().upper()
    if not ma: continue
    with_ma += 1
    if '.' in ma: with_dot += 1
    base = ma.split('.')[0]
    parents[base].append(ma)

three_with_children = sum(1 for k,v in parents.items() if len(v)>1 and any('.' in x for x in v))
print('\nStats: rows', total, 'with_ma', with_ma, 'with_dot', with_dot)
print('3-char groups with subcodes:', three_with_children)
