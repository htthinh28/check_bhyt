#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Đọc Excel danh mục chống chỉ định thuốc (BV) → sinh ma_nguon/tien_ich/seed_icd_drug_contra_bhyt.json
để gộp vào ICD_DRUG_CONTRA trong dong_co_giam_dinh.jsx.

Định dạng A (legacy): Loai, Mã thuốc, Tên hoạt chất, tên thuốc, ICD-10 chống chỉ định, …

Định dạng B (export Mapping / Catalog_mapping*.xlsx, sheet «mapping»):
  Ma_nguon = ICD chống chỉ định; Ma_thuc_hien = mã thuốc M03; Ten_thuc_hien = tên thuốc (nhiều mã thì «;»);
  Loai = THUOC_CCD_xx hoặc ICD_DRUG_CONTRA.

Chạy:
  python scripts/build_seed_icd_drug_contra_from_excel.py
  python scripts/build_seed_icd_drug_contra_from_excel.py --input "C:\\path\\file.xlsx"
  python scripts/build_seed_icd_drug_contra_from_excel.py --input catalog.xlsx --merge --out ma_nguon/tien_ich/seed_icd_drug_contra_bhyt.json
"""

from __future__ import annotations

import argparse
import json
import re
import sys
import uuid
from datetime import datetime, timezone
from pathlib import Path

try:
    import openpyxl
except ImportError:
    print("Cần cài: pip install openpyxl", file=sys.stderr)
    sys.exit(1)

REPO_ROOT = Path(__file__).resolve().parents[1]
DEFAULT_OUT = REPO_ROOT / "ma_nguon" / "tien_ich" / "seed_icd_drug_contra_bhyt.json"
DEFAULT_INPUT = Path(
    r"c:\Users\admin\Documents\Google Drive\BHYT\danh muc benh vien\thuoc chong chỉ dinh rhuoc.xlsx"
)


def tach_nhieu_ma(s: str) -> list[str]:
    t = str(s or "").strip()
    if not t:
        return []
    for ch in "|":
        t = t.replace(ch, ";")
    parts = re.split(r"[;,]", t)
    out: list[str] = []
    for p in parts:
        x = str(p or "").strip()
        if x:
            out.append(x)
    return out


def chuan_loai_mapping(loai_raw: str) -> str:
    u = str(loai_raw or "").strip().upper().replace(" ", "_")
    if u.startswith("ICD_DRUG_CONTRA") or u.startswith("THUOC_CCD"):
        return "ICD_DRUG_CONTRA"
    return u


def build_row(
    *,
    loai: str,
    ma_thuoc_cell: str,
    ten_hoat_chat_cell: str,
    ten_thuoc_cell: str,
    icd_cell: str,
    ten_benh_cell: str,
    quy_tac_cell: str,
    canh_bao_cell: str,
    line_no: int,
) -> dict | None:
    icds = tach_nhieu_ma(icd_cell)
    mas = tach_nhieu_ma(ma_thuoc_cell)
    if not icds or not mas:
        return None
    ten_aliases = [x for x in tach_nhieu_ma(ten_thuoc_cell) if x]
    hoat_aliases = [x for x in tach_nhieu_ma(ten_hoat_chat_cell) if x]
    mt = chuan_loai_mapping(loai)
    if mt != "ICD_DRUG_CONTRA":
        return None
    rid = f"SEED_ICD_CD_BHYT_L{line_no}"
    ts = datetime.now(timezone.utc).isoformat()
    md = {
        "rule_id": rid,
        "source_icd_codes": icds,
        "target_codes": mas,
        "khop_bang_ten_hoat_chat": True,
        "ten_thuoc_aliases": ten_aliases,
        "hoat_chat_aliases": hoat_aliases,
        "ten_benh_chong_chi_dinh": str(ten_benh_cell or "").strip(),
        "quy_tac_giam_dinh_excel": str(quy_tac_cell or "").strip(),
        "canh_bao_excel": str(canh_bao_cell or "").strip(),
        "excel_line": line_no,
    }
    return {
        "id": str(uuid.uuid4()),
        "mapping_type": "ICD_DRUG_CONTRA",
        "source_catalog": "icd10",
        "target_catalog": "drug_items",
        "source_id": 0,
        "target_id": 0,
        "source_code": ";".join(icds),
        "target_code": ";".join(mas),
        "effective_from": None,
        "effective_to": None,
        "priority": 0,
        "is_active": True,
        "metadata": md,
        "approval_status": "APPROVED",
        "created_at": ts,
        "updated_at": ts,
        "created_by": "build_seed_icd_drug_contra_from_excel.py",
        "updated_by": "",
    }


def _split_semicolon_only(s: str) -> list[str]:
    """Chỉ tách theo `;` — giữ dấu phẩy trong tên thuốc (vd 0,10g/ml)."""
    t = str(s or "").strip()
    if not t:
        return []
    return [p.strip() for p in t.split(";") if p.strip()]


def _norm_icd_key(icds: list[str]) -> tuple[str, ...]:
    return tuple(sorted({str(x or "").strip().upper().replace(" ", "") for x in icds if str(x or "").strip()}))


def _merge_key_one_target(rec: dict) -> tuple[tuple[str, ...], str]:
    """Một bản ghi đơn mã thuốc (sau tách) → khóa dedupe."""
    keys = merge_keys_record(rec)
    if not keys:
        return (tuple(), "")
    return keys[0]


def merge_keys_record(rec: dict) -> list[tuple[tuple[str, ...], str]]:
    """Mọi cặp (tập ICD chuẩn hóa, mã thuốc) trong một bản ghi seed."""
    md = rec.get("metadata") or {}
    icds = md.get("source_icd_codes")
    if not isinstance(icds, list):
        icds = tach_nhieu_ma(rec.get("source_code") or "")
    tgts = md.get("target_codes")
    if not isinstance(tgts, list):
        tgts = tach_nhieu_ma(rec.get("target_code") or "")
    ik = _norm_icd_key(icds)
    return [(ik, str(t).strip()) for t in tgts if str(t).strip()]


def build_rows_catalog_mapping_row(
    *,
    loai: str,
    ma_nguon: str,
    ma_thuc_hien: str,
    ten_thuc_hien: str,
    ten_nguon: str,
    line_no: int,
) -> list[dict]:
    """Export Mapping nghiệp vụ: Ma_nguon=ICD chống chỉ định, Ma_thuc_hien=mã thuốc M03."""
    icds = tach_nhieu_ma(ma_nguon)
    codes = _split_semicolon_only(ma_thuc_hien)
    names = _split_semicolon_only(ten_thuc_hien)
    if not icds or not codes:
        return []
    if chuan_loai_mapping(loai) != "ICD_DRUG_CONTRA":
        return []
    out: list[dict] = []
    if len(names) == len(codes):
        pairs = list(zip(codes, names))
    elif len(names) == 0:
        pairs = [(c, "") for c in codes]
    elif len(names) == 1:
        pairs = [(c, names[0]) for c in codes]
    else:
        pairs = [(codes[i], names[i] if i < len(names) else "") for i in range(len(codes))]
    for i, (code, tname) in enumerate(pairs):
        rid = f"SEED_ICD_CD_CAT_L{line_no}_{i + 1}"
        ts = datetime.now(timezone.utc).isoformat()
        ten_aliases = [tname] if tname else []
        md = {
            "rule_id": rid,
            "source_icd_codes": icds,
            "target_codes": [code],
            "khop_bang_ten_hoat_chat": True,
            "ten_thuoc_aliases": ten_aliases,
            "hoat_chat_aliases": [],
            "ten_benh_chong_chi_dinh": str(ten_nguon or "").strip(),
            "quy_tac_giam_dinh_excel": "",
            "canh_bao_excel": "",
            "excel_line": line_no,
            "catalog_export": True,
        }
        out.append(
            {
                "id": str(uuid.uuid4()),
                "mapping_type": "ICD_DRUG_CONTRA",
                "source_catalog": "icd10",
                "target_catalog": "drug_items",
                "source_id": 0,
                "target_id": 0,
                "source_code": ";".join(icds),
                "target_code": code,
                "effective_from": None,
                "effective_to": None,
                "priority": 0,
                "is_active": True,
                "metadata": md,
                "approval_status": "APPROVED",
                "created_at": ts,
                "updated_at": ts,
                "created_by": "build_seed_icd_drug_contra_from_excel.py",
                "updated_by": "",
            }
        )
    return out


def main() -> int:
    ap = argparse.ArgumentParser()
    ap.add_argument("--input", type=Path, default=DEFAULT_INPUT)
    ap.add_argument("--out", type=Path, default=DEFAULT_OUT)
    ap.add_argument(
        "--merge",
        action="store_true",
        help="Giữ các bản ghi seed sẵn có trong --out, chỉ thêm bản ghi mới (dedupe theo ICD+từng mã thuốc).",
    )
    args = ap.parse_args()
    inp: Path = args.input
    out: Path = args.out
    if not inp.is_file():
        print(f"Không tìm thấy file: {inp}", file=sys.stderr)
        return 1
    wb = openpyxl.load_workbook(inp, read_only=True, data_only=True)
    sheet_name = "mapping" if "mapping" in wb.sheetnames else wb.sheetnames[0]
    sh = wb[sheet_name]
    rows_iter = sh.iter_rows(min_row=1, values_only=True)
    header = next(rows_iter, None)
    if not header:
        print("Sheet trống.", file=sys.stderr)
        return 1
    hmap = {str(c or "").strip(): i for i, c in enumerate(header)}

    def col(*names: str) -> int | None:
        for n in names:
            if n in hmap:
                return hmap[n]
        return None

    idx_ma_nguon = col("Ma_nguon", "Mã nguồn")
    idx_ma_th = col("Ma_thuc_hien", "Mã TH", "target_code")
    idx_ten_th = col("Ten_thuc_hien", "Tên TH")
    idx_ten_nguon = col("Ten_nguon", "Tên nguồn")
    idx_loai_cat = col("Loai", "Loại")

    idx_loai = col("Loai", "Loại")
    idx_ma = col("Mã thuốc")
    idx_hoat = col("Tên hoạt chất")
    idx_ten_thuoc = col("tên thuốc ", "tên thuốc", "Tên thuốc")
    idx_icd = col("ICD-10 chống chỉ định")
    idx_ten_benh = col("Tên bệnh chống chỉ định")
    idx_qt = col("Quy tắc kiểm tra")
    idx_cb = col("Cảnh báo")

    use_catalog = idx_ma_nguon is not None and idx_ma_th is not None and idx_icd is None
    if not use_catalog:
        need = [idx_loai, idx_ma, idx_icd]
        if any(x is None for x in need):
            print(f"Thiếu cột bắt buộc trong header: {header}", file=sys.stderr)
            return 1

    existing: list[dict] = []
    seen: set[tuple[tuple[str, ...], str]] = set()
    if args.merge and out.is_file():
        try:
            existing = json.loads(out.read_text(encoding="utf-8"))
            if not isinstance(existing, list):
                existing = []
        except json.JSONDecodeError:
            existing = []
        for rec in existing:
            for k in merge_keys_record(rec):
                seen.add(k)

    out_rows: list[dict] = []
    line = 1
    for tup in rows_iter:
        line += 1
        if not tup or all(v is None or str(v).strip() == "" for v in tup):
            continue

        def gv(i: int | None) -> str:
            if i is None or i >= len(tup):
                return ""
            v = tup[i]
            return "" if v is None else str(v).strip()

        if use_catalog:
            loai_c = gv(idx_loai_cat)
            mn = gv(idx_ma_nguon)
            mth = gv(idx_ma_th)
            if not loai_c or not mn or not mth:
                continue
            for r in build_rows_catalog_mapping_row(
                loai=loai_c,
                ma_nguon=mn,
                ma_thuc_hien=mth,
                ten_thuc_hien=gv(idx_ten_th),
                ten_nguon=gv(idx_ten_nguon),
                line_no=line,
            ):
                k = _merge_key_one_target(r)
                if k in seen:
                    continue
                seen.add(k)
                out_rows.append(r)
            continue

        loai = gv(idx_loai)
        if not loai:
            continue
        r = build_row(
            loai=loai,
            ma_thuoc_cell=gv(idx_ma),
            ten_hoat_chat_cell=gv(idx_hoat),
            ten_thuoc_cell=gv(idx_ten_thuoc),
            icd_cell=gv(idx_icd),
            ten_benh_cell=gv(idx_ten_benh),
            quy_tac_cell=gv(idx_qt),
            canh_bao_cell=gv(idx_cb),
            line_no=line,
        )
        if r:
            if args.merge:
                keys = merge_keys_record(r)
                if keys and all(k not in seen for k in keys):
                    for k in keys:
                        seen.add(k)
                    out_rows.append(r)
            else:
                out_rows.append(r)

    final_rows = (existing + out_rows) if args.merge else out_rows

    out.parent.mkdir(parents=True, exist_ok=True)
    with out.open("w", encoding="utf-8") as f:
        json.dump(final_rows, f, ensure_ascii=False, indent=2)
    mode = "merge" if args.merge else "replace"
    sheet_info = f"sheet={sheet_name}"
    print(f"[{mode}] +{len(out_rows)} new, total {len(final_rows)} -> {out} ({sheet_info})")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
