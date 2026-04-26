#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Danh mục BV (sheet Sheet1, cột LOAI_QUY_TAC):
  - Đổi LOAI_QUY_TAC từ DATASET → BUILTIN cho mọi dòng đang DATASET
    (đồng bộ ý nghĩa với bundle seed trong app — ma_nguon/tien_ich/du_lieu_luat_pttt_muc11.jsx).
  - Đồng bộ một số MA_LUAT với nội dung seed (mặc định: DVKT_1742) khi catalog BV lệch bản bundle.

Chạy:
  python scripts/promote_bv_catalog_pttt_dataset_to_builtin.py
  python scripts/promote_bv_catalog_pttt_dataset_to_builtin.py --input "C:\\path\\file.xlsx" --dry-run
  python scripts/promote_bv_catalog_pttt_dataset_to_builtin.py --output "C:\\path\\out_builtin.xlsx"
    (khi file nguồn bị khóa — đóng Excel/Google sync rồi thay thế tay hoặc chạy lại không --output)
"""

from __future__ import annotations

import argparse
import ast
import re
import sys
from pathlib import Path

try:
    import openpyxl
except ImportError:
    print("Cần cài: pip install openpyxl", file=sys.stderr)
    sys.exit(1)

REPO_ROOT = Path(__file__).resolve().parents[1]
DEFAULT_INPUT = Path(
    r"c:\Users\admin\Documents\Google Drive\BHYT\danh muc benh vien\danh muc quy tac phau thua thu thu.xlsx"
)
SEED_JSX = REPO_ROOT / "ma_nguon" / "tien_ich" / "du_lieu_luat_pttt_muc11.jsx"


def load_seed_by_ma() -> dict[str, dict]:
    text = SEED_JSX.read_text(encoding="utf-8")
    m = re.search(r"export const DU_LIEU_SEED_LUAT_PTTT_MUC11 = (\[[\s\S]*?\]);", text)
    if not m:
        raise ValueError("Không đọc được DU_LIEU_SEED_LUAT_PTTT_MUC11 từ du_lieu_luat_pttt_muc11.jsx")
    rows = ast.literal_eval(m.group(1))
    out: dict[str, dict] = {}
    for r in rows:
        ma = str(r.get("MA_LUAT", "")).strip().upper()
        if ma:
            out[ma] = r
    return out


def main() -> None:
    ap = argparse.ArgumentParser()
    ap.add_argument("--input", type=Path, default=DEFAULT_INPUT)
    ap.add_argument(
        "--output",
        type=Path,
        default=None,
        help="File .xlsx đích (mặc định: ghi đè --input)",
    )
    ap.add_argument("--dry-run", action="store_true", help="Chỉ in thống kê, không ghi file")
    ap.add_argument(
        "--sync-ma-from-seed",
        default="DVKT_1742",
        help="Danh sách MA (phẩy) đồng bộ TEN/DIEU_KIEN/CANH_BAO từ seed vào Excel",
    )
    args = ap.parse_args()
    if not args.input.exists():
        print(f"Không thấy file: {args.input}", file=sys.stderr)
        sys.exit(1)

    seed_by_ma = load_seed_by_ma()
    sync_mas = {x.strip().upper() for x in str(args.sync_ma_from_seed).split(",") if x.strip()}

    wb = openpyxl.load_workbook(args.input, data_only=False)
    if "Sheet1" not in wb.sheetnames:
        print(f"Không có Sheet1. Có: {wb.sheetnames}", file=sys.stderr)
        sys.exit(1)
    ws = wb["Sheet1"]
    header_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True))
    headers = [str(c or "").strip() for c in header_row]
    col = {h.upper(): i + 1 for i, h in enumerate(headers)}  # 1-based

    required = ["LOAI_QUY_TAC", "MA_LUAT", "TEN_QUY_TAC", "DIEU_KIEN", "CANH_BAO", "TRANG_THAI"]
    for h in required:
        if h not in col:
            print(f"Thiếu cột {h}. Có: {headers}", file=sys.stderr)
            sys.exit(1)

    chi_col = col.get("CHI_TIET_CANH_BAO")

    n_promote = 0
    n_sync = 0
    for r in range(2, ws.max_row + 1):
        loai = ws.cell(r, col["LOAI_QUY_TAC"]).value
        loai_u = str(loai or "").strip().upper()
        if loai_u == "DATASET":
            n_promote += 1
            if not args.dry_run:
                ws.cell(r, col["LOAI_QUY_TAC"]).value = "BUILTIN"

        ma = str(ws.cell(r, col["MA_LUAT"]).value or "").strip().upper()
        if ma in sync_mas and ma in seed_by_ma:
            s = seed_by_ma[ma]
            if not args.dry_run:
                ws.cell(r, col["TEN_QUY_TAC"]).value = s.get("TEN_QUY_TAC") or ""
                ws.cell(r, col["DIEU_KIEN"]).value = s.get("DIEU_KIEN") or ""
                ws.cell(r, col["CANH_BAO"]).value = s.get("CANH_BAO") or ""
                ws.cell(r, col["TRANG_THAI"]).value = str(s.get("TRANG_THAI") or "ON").strip().upper() or "ON"
                if chi_col:
                    ws.cell(r, chi_col).value = s.get("CANH_BAO") or ""
            n_sync += 1

    print(f"DATASET → BUILTIN: {n_promote} dòng")
    print(f"Đồng bộ từ seed ({', '.join(sorted(sync_mas))}): {n_sync} dòng ghi (theo MA_LUAT)")
    if args.dry_run:
        print("(dry-run, không lưu file)")
    else:
        out = args.output or args.input
        out.parent.mkdir(parents=True, exist_ok=True)
        wb.save(out)
        print(f"Đã lưu: {out}")


if __name__ == "__main__":
    main()
